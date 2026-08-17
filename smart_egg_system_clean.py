import csv
import os
import matplotlib.pyplot as plt
import pandas as pd
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Image
from datetime import datetime
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image
)

import getpass


def load_users():
    users = []

    users_path = os.path.join(os.path.dirname(__file__), "users.csv")

    try:
        with open(users_path, "r", newline="", encoding="utf-8") as file:
            reader = csv.DictReader(file, delimiter=";")
            users = list(reader)

    except FileNotFoundError:
        print("users.csv not found!")

    return users
def save_users(users):

    users_path = os.path.join(os.path.dirname(__file__), "users.csv")

    with open(users_path, "w", newline="", encoding="utf-8") as file:

        fieldnames = ["username", "password", "role"]

        writer = csv.DictWriter(file,
                                fieldnames=fieldnames,
                                delimiter=";")

        writer.writeheader()
        writer.writerows(users)

def login():

    users = load_users()

    max_attempts = 3

    print("=" * 40)
    print(" SMART EGG PRODUCTION SYSTEM")
    print("=" * 40)

    for attempt in range(max_attempts):

        username = input("Username: ")
        password = getpass.getpass("Password: ")

        for user in users:

            if (username == user["username"] and
                    password == user["password"]):

                print(f"\nWelcome {username}!")
                print(f"Role: {user['role']}")
                print("Login successful!\n")

                return user

        remaining = max_attempts - attempt - 1

        if remaining > 0:
            print(f"\nInvalid username or password.")
            print(f"Attempts remaining: {remaining}\n")

    print("\nToo many failed login attempts.")
    return None

def user_management(current_user):

    while True:
        print("\n===== USER MANAGEMENT =====")
        print("1. Add User")
        print("2. View Users")
        print("3. Change Password")
        print("4. Delete User")
        print("5. Back")

        choice = input("Choose an option: ")

        if choice == "1":
            add_user(current_user)

        elif choice == "2":
            view_users(current_user)

        elif choice == "3":
            change_password(current_user)

        elif choice == "4":
            delete_user(current_user)

        elif choice == "5":
            break

        else:
            print("Invalid choice!")
            
def add_user(current_user):

    if current_user["role"] != "Administrator":
        print("Access denied! Only Administrators can add users.")
        return

    print("\n=== ADD USER ===")

    username = input("Username: ")
    password = input("Password: ")
    role = input("Role (Administrator/Manager/Employee): ")

    users = load_users()

    users.append({
        "username": username,
        "password": password,
        "role": role
    })

    save_users(users)

    print("User added successfully!")
    
    
def view_users(current_user):

    if current_user["role"] not in ["Administrator", "Manager"]:
        print("Access denied!")
        return

    users = load_users()

    print("\n===== USERS =====")

    for user in users:
        print(f"{user['username']} - {user['role']}") 
   
   
def change_password(current_user):

    users = load_users()

    new_password = input("Enter new password: ")

    for user in users:

        if user["username"] == current_user["username"]:
            user["password"] = new_password

    save_users(users)

    print("Password changed successfully!")
    
    
def delete_user(current_user):

    if current_user["role"] != "Administrator":
        print("Access denied! Only Administrators can delete users.")
        return

    users = load_users()

    username = input("Username to delete: ")

    users = [user for user in users if user["username"] != username]

    save_users(users)

    print("User deleted successfully!")       

def load_records():
    global records

    import os

    csv_path = os.path.join(os.path.dirname(__file__), "records.csv")

    print("Loading:", csv_path)
    print("Exists:", os.path.exists(csv_path))

    try:
        with open(csv_path, "r", newline="", encoding="utf-8") as file:
            reader = csv.DictReader(file, delimiter=";")
            records = list(reader)

        # Convert numeric fields
        for r in records:
            r["layers"] = int(r["layers"])
            r["eggs_produced"] = int(r["eggs_produced"])
            r["feed_used"] = float(r["feed_used"])
            r["feed_price"] = float(r["feed_price"])
            r["feed_cost"] = float(r["feed_cost"])
            r["eggs_sold"] = int(r["eggs_sold"])
            r["price_per_egg"] = float(r["price_per_egg"])
            r["revenue"] = float(r["revenue"])

        print("Loaded", len(records), "records.")

    except Exception as e:
        print("Error loading records:", e)
        records = []
        

def save_records():
    csv_path = os.path.join(os.path.dirname(__file__), "records.csv")

    fieldnames = [
        "date",
        "layers",
        "eggs_produced",
        "feed_used",
        "feed_price",
        "feed_cost",
        "eggs_sold",
        "price_per_egg",
        "revenue",
        "notes"
    ]

    try:
        with open(csv_path, "w", newline="", encoding="utf-8") as file:
            writer = csv.DictWriter(file, fieldnames=fieldnames, delimiter=";")
            writer.writeheader()
            writer.writerows(records)

    except Exception as e:
        print("Error saving records:", e)
        
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference

def export_to_excel():
    if not records:
        print("\nNo records to export.")
        return

    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, Reference

        # Create DataFrame
        df = pd.DataFrame(records)

        filename = "Egg_Production_Report.xlsx"

        # Export records to Excel
        df.to_excel(filename, index=False)

        # Open workbook
        wb = load_workbook(filename)

        # ==========================
        # SHEET 1 - Egg Production
        # ==========================

        ws = wb.active
        ws.title = "Egg Production"

        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                fill_type="solid",
                start_color="4F81BD",
                end_color="4F81BD"
            )
            cell.alignment = Alignment(horizontal="center")

        # Auto-fit columns
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 3

        # Currency formatting
        for col in ["F", "I"]:
            for cell in ws[col][1:]:
                cell.number_format = '"R"#,##0.00'

        # ==========================
        # SHEET 2 - Summary
        # ==========================

        summary_ws = wb.create_sheet(title="Summary")

        # Title
        summary_ws["A1"] = "SMART EGG PRODUCTION REPORT"
        summary_ws.merge_cells("A1:B1")

        summary_ws["A1"].font = Font(
            bold=True,
            size=18,
            color="FFFFFF"
        )

        summary_ws["A1"].fill = PatternFill(
            fill_type="solid",
            start_color="2E75B6",
            end_color="2E75B6"
        )

        summary_ws["A1"].alignment = Alignment(horizontal="center")

        # Headers
        summary_ws["A3"] = "Metric"
        summary_ws["B3"] = "Value"

        for cell in ["A3", "B3"]:
            summary_ws[cell].font = Font(bold=True, color="FFFFFF")
            summary_ws[cell].fill = PatternFill(
                fill_type="solid",
                start_color="4F81BD",
                end_color="4F81BD"
            )
            summary_ws[cell].alignment = Alignment(horizontal="center")

        # Totals
        total_records = len(records)
        total_layers = sum(int(r["layers"]) for r in records)
        total_eggs_produced = sum(int(r["eggs_produced"]) for r in records)
        total_eggs_sold = sum(int(r["eggs_sold"]) for r in records)
        total_feed_used = sum(float(r["feed_used"]) for r in records)
        total_feed_cost = sum(float(r["feed_cost"]) for r in records)
        total_revenue = sum(float(r["revenue"]) for r in records)

        average_production_rate = (
            (total_eggs_produced / total_layers) * 100
            if total_layers > 0 else 0
        )

        summary_data = [
            ("Total Records", total_records),
            ("Total Layers", total_layers),
            ("Total Eggs Produced", total_eggs_produced),
            ("Total Eggs Sold", total_eggs_sold),
            ("Total Feed Used (kg)", total_feed_used),
            ("Total Feed Cost (R)", total_feed_cost),
            ("Total Revenue (R)", total_revenue),
            ("Average Production Rate (%)", round(average_production_rate, 2))
        ]

        # Write summary table
        for row, (metric, value) in enumerate(summary_data, start=4):
            summary_ws[f"A{row}"] = metric
            summary_ws[f"B{row}"] = value

        # Currency formatting
        summary_ws["B9"].number_format = '"R"#,##0.00'
        summary_ws["B10"].number_format = '"R"#,##0.00'

        # Column widths
        summary_ws.column_dimensions["A"].width = 35
        summary_ws.column_dimensions["B"].width = 20

        # ==========================
        # Performance Chart Data
        # ==========================

        summary_ws["D3"] = "Metric"
        summary_ws["E3"] = "Value"

        summary_ws["D4"] = "Eggs Produced"
        summary_ws["E4"] = total_eggs_produced

        summary_ws["D5"] = "Eggs Sold"
        summary_ws["E5"] = total_eggs_sold

        summary_ws["D6"] = "Feed Used (kg)"
        summary_ws["E6"] = total_feed_used

        # Create Bar Chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Farm Performance"
        chart.y_axis.title = "Quantity"

        data = Reference(summary_ws,
                         min_col=5,
                         min_row=3,
                         max_row=6)

        categories = Reference(summary_ws,
                               min_col=4,
                               min_row=4,
                               max_row=6)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        chart.width = 14
        chart.height = 8

        summary_ws.add_chart(chart, "D8")

        # Save workbook
        wb.save(filename)

        print(f"\n✅ Excel report exported successfully as '{filename}'")

    except Exception as e:
        print("Error exporting report:", e)
        

from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image
)

def add_page_footer(canvas, doc):

    canvas.saveState()

    width, height = doc.pagesize

    canvas.setStrokeColor(colors.grey)

    canvas.line(
        doc.leftMargin,
        55,
        width-doc.rightMargin,
        55
    )

    canvas.setFont("Helvetica",8)

    canvas.drawString(
        doc.leftMargin,
        40,
        "Smart Egg Production System © 2026 | Developed by Ntombizodwa Mnisi"
    )

    canvas.drawRightString(
        width-doc.rightMargin,
        40,
        f"Page {canvas.getPageNumber()}"
    )

    canvas.restoreState()

def export_to_pdf():

    if not records:
        print("\nNo records to export.")
        return

    try:

        pdf = SimpleDocTemplate(
            "Egg_Production_Report.pdf",
            rightMargin=40,
            leftMargin=40,
            topMargin=50,
            bottomMargin=60
        )

        styles = getSampleStyleSheet()

        elements = []

        ##################################################
        # TITLE
        ##################################################

        elements.append(
            Paragraph(
                "SMART EGG PRODUCTION REPORT",
        styles["Title"]
    )
)

        elements.append(Spacer(1, 15))

        report_date = datetime.now().strftime("%d %B %Y  %H:%M:%S")

        elements.append(
            Paragraph(
        f"<b>Report Generated:</b> {report_date}",
        styles["Heading3"]
    )
)

        elements.append(
    Paragraph(
        "<b>Prepared By:</b> Ntombizodwa Mnisi",
        styles["Heading3"]
    )
)

        elements.append(Spacer(1,25))

        ##################################################
        # CALCULATIONS
        ##################################################

        total_records = len(records)

        total_layers = sum(
            int(r["layers"])
            for r in records
        )

        total_eggs_produced = sum(
            int(r["eggs_produced"])
            for r in records
        )

        total_eggs_sold = sum(
            int(r["eggs_sold"])
            for r in records
        )

        total_feed_used = sum(
            float(r["feed_used"])
            for r in records
        )

        total_feed_cost = sum(
            float(r.get("feed_cost", 0))
            for r in records
        )

        total_revenue = sum(
            float(r["revenue"])
            for r in records
        )

        production_rate = (
            (total_eggs_produced / total_layers) * 100
            if total_layers else 0
        )

        ##################################################
        # SUMMARY
        ##################################################

        elements.append(
            Paragraph(
                "<b>Farm Performance Summary</b>",
                styles["Heading2"]
            )
        )

        elements.append(Spacer(1, 8))

        summary = [

            ["Metric", "Value"],

            ["Total Records", total_records],

            ["Total Layers", total_layers],

            ["Eggs Produced", total_eggs_produced],

            ["Eggs Sold", total_eggs_sold],

            ["Feed Used (kg)", f"{total_feed_used:.2f}"],

            ["Feed Cost", f"R {total_feed_cost:,.2f}"],

            ["Revenue", f"R {total_revenue:,.2f}"],

            ["Production Rate", f"{production_rate:.2f}%"]
        ]

        table = Table(summary)

        table.setStyle(TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.darkblue),

            ("TEXTCOLOR",(0,0),(-1,0),colors.white),

            ("GRID",(0,0),(-1,-1),1,colors.black),

            ("BACKGROUND",(0,1),(-1,-1),colors.whitesmoke),

            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

            ("ALIGN",(1,1),(-1,-1),"RIGHT")

        ]))

        elements.append(table)

        elements.append(Spacer(1,20))

        ##################################################
        # CHART
        ##################################################

        plt.figure(figsize=(6,4))

        plt.bar(

            ["Produced","Sold","Feed"],

            [

                total_eggs_produced,

                total_eggs_sold,

                total_feed_used

            ]

        )

        plt.title("Farm Performance")

        plt.tight_layout()

        plt.savefig("farm_chart.png")

        plt.close()

        chart = Image("farm_chart.png")

        chart.drawWidth = 5*inch

        chart.drawHeight = 3*inch

        elements.append(chart)

        elements.append(Spacer(1,20))

        ##################################################
        # RECORDS
        ##################################################

        elements.append(
            Paragraph(
                "<b>Daily Production Records</b>",
                styles["Heading2"]
            )
        )

        data = [[

            "Date",

            "Layers",

            "Eggs",

            "Sold",

            "Feed",

            "Revenue"

        ]]

        for r in records:

            data.append([

                r["date"],

                r["layers"],

                r["eggs_produced"],

                r["eggs_sold"],

                f"{float(r['feed_used']):.2f}",

                f"R {float(r['revenue']):,.2f}"

            ])

        record_table = Table(
    data,
    repeatRows=1
)

        record_table.setStyle(TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.darkblue),

            ("TEXTCOLOR",(0,0),(-1,0),colors.white),

            ("GRID",(0,0),(-1,-1),1,colors.black),

            ("BACKGROUND",(0,1),(-1,-1),colors.beige),

            ("ALIGN",(0,0),(-1,-1),"CENTER"),

            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold")

        ]))

        elements.append(record_table)

        ##################################################
        # BUILD PDF
        ##################################################

        pdf.build(

            elements,

            onFirstPage=add_page_footer,

            onLaterPages=add_page_footer

        )

        print(
            "\n✅ PDF exported successfully as Egg_Production_Report.pdf"
        )

    except Exception as e:

        print("Error exporting PDF:", e)
# ---------------- DISPLAY HELPER ----------------

def calculate_production_rate(layers, eggs_produced):
    layers = int(layers)
    eggs_produced = int(eggs_produced)

    if layers == 0:
        return 0

    return (eggs_produced / layers) * 100


def print_record(record):
    rate = calculate_production_rate(
        record["layers"],
        record["eggs_produced"]
    )

    print("\n----------------------")
    print("Date:", record["date"])
    print("Layers:", record["layers"])
    print("Eggs Produced:", record["eggs_produced"])
    print("Eggs Sold:", record["eggs_sold"])
    print("Feed Used:", record["feed_used"], "kg")
    print(f"Revenue: R{record['revenue']:.2f}")
    print(f"Production Rate: {rate:.2f}%")
    print("Notes:", record["notes"])


# ---------------- CORE FUNCTIONS ----------------

def add_record():
    print("\n--- ADD RECORD ---")

    try:
        # Collect input
        date = input("Enter date (YYYY-MM-DD): ")

        for record in records:
            if record["date"] == date:
                print("❌ A record for this date already exists.")
                return
            
        layers = int(input("Number of layers: "))
        eggs_produced = int(input("Eggs produced: "))
        feed_used = float(input("Feed used (kg): "))
        feed_price = float(input("Feed price per kg (R): "))
        eggs_sold = int(input("Eggs sold: "))
        price_per_egg = float(input("Price per egg (R): "))
        notes = input("Notes (optional): ")

        # Validation
        if layers <= 0:
            print("Number of layers must be greater than 0.")
            return

        if eggs_produced < 0 or eggs_sold < 0:
            print("Egg quantities cannot be negative.")
            return

        if feed_used < 0 or feed_price < 0 or price_per_egg < 0:
            print("Feed and price values cannot be negative.")
            return

        if eggs_sold > eggs_produced:
            print("Eggs sold cannot exceed eggs produced.")
            return
        
    
        feed_cost = feed_used * feed_price
        revenue = eggs_sold * price_per_egg
        production_rate = calculate_production_rate(layers, eggs_produced)

        # Create record
        record = {
            "date": date,
            "layers": layers,
            "eggs_produced": eggs_produced,
            "feed_used": feed_used,
            "feed_price": feed_price,
            "feed_cost": feed_cost,
            "eggs_sold": eggs_sold,
            "price_per_egg": price_per_egg,
            "revenue": revenue,
            "notes": notes
        }

        # Save record
        records.append(record)
        save_records()

        # Display results
        print("\nRecord added successfully.")
        print(f"Revenue: R{revenue:.2f}")
        print(f"Feed Cost: R{feed_cost:.2f}")
        print(f"Production Rate: {production_rate:.2f}%")

        if production_rate < 70:
            print("Warning: Production is below 70%. Check feed, water, weather, stress, and flock health.")
        else:
            print("Production level is within a normal range.")

    except ValueError:
        print("Invalid input. Please enter the correct numeric values.")
def view_records():
    print("\n--- VIEW RECORDS ---")

    if not records:
        print("No records found.")
        return

    for record in records:
        print_record(record)


def summary():
    print("\n--- SUMMARY ---")

    if not records:
        print("No data available.")
        return

    total_eggs = sum(record["eggs_produced"] for record in records)
    total_sold = sum(record["eggs_sold"] for record in records)
    total_revenue = sum(record["revenue"] for record in records)

    print("Total Eggs Produced:", total_eggs)
    print("Total Eggs Sold:", total_sold)
    print(f"Total Revenue: R{total_revenue:.2f}")


def search_record():
    print("\n--- SEARCH RECORD ---")

    date = input("Enter date (YYYY-MM-DD): ")

    for record in records:
        if record["date"] == date:
            print("\n--- RECORD FOUND ---")
            print_record(record)
            return

    print("No record found.")


def delete_record():
    print("\n--- DELETE RECORD ---")

    date = input("Enter date to delete: ")

    for record in records:
        if record["date"] == date:
            records.remove(record)
            save_records()
            print("Record deleted successfully.")
            return

    print("Record not found.")


def edit_record():
    print("\n--- EDIT RECORD ---")

    date = input("Enter date to edit: ")

    for record in records:
        if record["date"] == date:
            try:
                record["layers"] = int(input("New number of layers: "))
                record["eggs_produced"] = int(input("New eggs produced: "))
                record["feed_used"] = float(input("New feed used (kg): "))
                record["eggs_sold"] = int(input("New eggs sold: "))
                record["revenue"] = float(input("New revenue: "))
                record["notes"] = input("New notes: ")

                if record["layers"] <= 0:
                    print("Number of layers must be greater than 0.")
                    return

                save_records()
                print("Record updated successfully.")

            except ValueError:
                print("Invalid input. Record was not updated.")

            return

    print("Record not found.")


# ---------------- CHARTS ---------------
import matplotlib.pyplot as plt
from collections import defaultdict
from datetime import datetime

def production_trend():
    print("\n--- EGG PRODUCTION TREND ---")

    if not records:
        print("No data to display.")
        return

    daily_totals = defaultdict(int)

    for record in records:
        daily_totals[record["date"]] += int(record["eggs_produced"])

    sorted_dates = sorted(
        daily_totals.keys(),
        key=lambda d: datetime.strptime(d, "%d %m %Y")
    )

    eggs = [daily_totals[d] for d in sorted_dates]

    plt.figure(figsize=(12, 6))
    plt.plot(sorted_dates, eggs, marker="o")
    plt.title("Egg Production Trend")
    plt.xlabel("Date")
    plt.ylabel("Eggs Produced")
    plt.xticks(rotation=45)
    plt.grid(True)
    plt.tight_layout()
    plt.show()

def production_rate_chart():
    print("\n--- PRODUCTION RATE TREND ---")

    if not records:
        print("No data to display.")
        return

    dates = [record["date"] for record in records]

    rates = [
        calculate_production_rate(
            record["layers"],
            record["eggs_produced"]
        )
        for record in records
    ]
    print("\n=== DEBUG ===")
    print("Number of records:", len(records))
    print("Dates:", dates)
    print("Rates:", rates)

    plt.figure(figsize=(12, 5))
    plt.plot(dates, rates, marker="o")
    plt.title("Production Rate Trend")
    plt.xlabel("Date")
    plt.ylabel("Production Rate (%)")
    plt.xticks(rotation=45)
    plt.grid(True)
    plt.tight_layout()
    plt.show()


# ---------------- DASHBOARD ----------------

def show_kpis():
    print("\n=== FARM DASHBOARD KPIs ===")

    if not records:
        print("No data available.")
        return

    total_eggs = sum(record["eggs_produced"] for record in records)
    total_sold = sum(record["eggs_sold"] for record in records)
    total_revenue = sum(record["revenue"] for record in records)
    total_layers = sum(record["layers"] for record in records)

    if total_layers > 0:
        average_rate = (total_eggs / total_layers) * 100
    else:
        average_rate = 0

    print("Total Eggs Produced:", total_eggs)
    print("Total Eggs Sold:", total_sold)
    print(f"Total Revenue: R{total_revenue:.2f}")
    print(f"Average Production Rate: {average_rate:.2f}%")

def best_worst_days():
    print("\n--- BEST & WORST PRODUCTION DAYS ---")

    if not records:
        print("No records available.")
        return

    best = None
    worst = None

    for record in records:
        if record["layers"] == 0:
            continue

        rate = (record["eggs_produced"] / record["layers"]) * 100

        if best is None or rate > best["rate"]:
            best = {
                "date": record["date"],
                "eggs": record["eggs_produced"],
                "layers": record["layers"],
                "rate": rate
            }

        if worst is None or rate < worst["rate"]:
            worst = {
                "date": record["date"],
                "eggs": record["eggs_produced"],
                "layers": record["layers"],
                "rate": rate
            }

    print("\n🥇 Best Production Day")
    print(f"Date: {best['date']}")
    print(f"Eggs Produced: {best['eggs']}")
    print(f"Layers: {best['layers']}")
    print(f"Production Rate: {best['rate']:.2f}%")

    print("\n🐔 Worst Production Day")
    print(f"Date: {worst['date']}")
    print(f"Eggs Produced: {worst['eggs']}")
    print(f"Layers: {worst['layers']}")
    print(f"Production Rate: {worst['rate']:.2f}%")


def best_worst_chart():
    print("\n--- BEST & WORST PRODUCTION DAYS CHART ---")

    if not records:
        print("No records available.")
        return

    rates = []

    for record in records:
        if record["layers"] > 0:
            rate = (record["eggs_produced"] / record["layers"]) * 100
            rates.append({
                "date": record["date"],
                "rate": rate
            })

    # Find best and worst
    best = max(rates, key=lambda x: x["rate"])
    worst = min(rates, key=lambda x: x["rate"])

    # Plot
    plt.figure(figsize=(6,5))
    plt.bar(
        ["Best Day", "Worst Day"],
        [best["rate"], worst["rate"]]
    )

    plt.title("Best vs Worst Production Days")
    plt.ylabel("Production Rate (%)")
    plt.ylim(0, 110)

    # Show values above bars
    plt.text(0, best["rate"] + 1,
             f"{best['rate']:.2f}%\n{best['date']}",
             ha="center")

    plt.text(1, worst["rate"] + 1,
             f"{worst['rate']:.2f}%\n{worst['date']}",
             ha="center")

    plt.show()



def daily_profit_analysis():
    print("\n===== DAILY PROFIT ANALYSIS =====")

    if len(records) == 0:
        print("No records available.")
        return

    total_profit = 0

    print(f"{'Date':<15}{'Revenue':<12}{'Feed Cost':<12}{'Profit'}")
    print("-" * 55)

    for r in records:
        revenue = r.get("revenue", 0)
        feed_cost = r.get("feed_cost", 0)

        profit = revenue - feed_cost
        total_profit += profit

        print(
            f"{r['date']:<15}"
            f"R{revenue:<11.2f}"
            f"R{feed_cost:<11.2f}"
            f"R{profit:.2f}"
        )

    print("-" * 55)
    print(f"Total Profit: R{total_profit:.2f}")

def production_efficiency():
    print("\n===== PRODUCTION EFFICIENCY =====")

    if len(records) == 0:
        print("No data available.")
        return

    total_produced = 0
    total_sold = 0

    for r in records:
        total_produced += r.get("eggs_produced", 0)
        total_sold += r.get("eggs_sold", 0)

    if total_produced == 0:
        print("No production data.")
        return

    efficiency = (total_sold / total_produced) * 100

    print(f"Total Eggs Produced: {total_produced}")
    print(f"Total Eggs Sold: {total_sold}")
    print(f"Efficiency: {efficiency:.2f}%")    

def monthly_revenue():
    print("\n===== MONTHLY REVENUE =====")

    if not records:
        print("No data available.")
        return

    monthly = {}

    for r in records:
        month = r["date"][:7]      # YYYY-MM
        monthly[month] = monthly.get(month, 0) + r["revenue"]

    print(f"{'Month':<12}{'Revenue'}")
    print("-" * 25)

    for month, revenue in sorted(monthly.items()):
        print(f"{month:<12}R{revenue:.2f}")

def monthly_production():
    print("\n===== MONTHLY EGG PRODUCTION =====")

    if not records:
        print("No data available.")
        return

    monthly = {}

    for r in records:
        month = r["date"][:7]
        monthly[month] = monthly.get(month, 0) + r["eggs_produced"]

    print(f"{'Month':<12}{'Eggs Produced'}")
    print("-" * 30)

    for month, eggs in sorted(monthly.items()):
        print(f"{month:<12}{eggs}")

def average_eggs():
    print("\n===== AVERAGE EGGS PER DAY =====")

    if not records:
        print("No data available.")
        return

    total = sum(r["eggs_produced"] for r in records)
    average = total / len(records)

    print(f"Average Eggs Produced Per Day: {average:.2f}")

def average_revenue():
    print("\n===== AVERAGE REVENUE PER DAY =====")

    if not records:
        print("No data available.")
        return

    total = sum(r["revenue"] for r in records)
    average = total / len(records)

    print(f"Average Revenue Per Day: R{average:.2f}")


def top_production_days():
    print("\n===== TOP 5 PRODUCTION DAYS =====")

    if not records:
        print("No data available.")
        return

    sorted_records = sorted(
        records,
        key=lambda r: r["eggs_produced"],
        reverse=True
    )

    print(f"{'Date':<15}{'Eggs Produced'}")
    print("-" * 30)

    for r in sorted_records[:5]:
        print(f"{r['date']:<15}{r['eggs_produced']}")

def worst_production_days():
    print("\n===== WORST 5 PRODUCTION DAYS =====")

    if not records:
        print("No data available.")
        return

    sorted_records = sorted(
        records,
        key=lambda r: r["eggs_produced"]
    )

    print(f"{'Date':<15}{'Eggs Produced'}")
    print("-" * 30)

    for r in sorted_records[:5]:
        print(f"{r['date']:<15}{r['eggs_produced']}")  


def business_analytics():
    while True:
        print("\n===== BUSINESS ANALYTICS =====")
        print("1. Daily Profit Analysis")
        print("2. Monthly Revenue")
        print("3. Monthly Egg Production")
        print("4. Average Eggs Per Day")
        print("5. Average Revenue Per Day")
        print("6. Production Efficiency")
        print("7. Top 5 Production Days")
        print("8. Worst 5 Production Days")
        print("9. Production Trend Graph")
        print("10. Return to Main Menu")

        choice = input("Choose an option: ")

        if choice == "1":
            daily_profit_analysis()

        elif choice == "2":
            monthly_revenue()

        elif choice == "3":
            monthly_production()

        elif choice == "4":
            average_eggs()

        elif choice == "5":
            average_revenue()

        elif choice == "6":
            production_efficiency()

        elif choice == "7":
            top_production_days()

        elif choice == "8":
            worst_production_days()

        elif choice == "9":
            production_trend()     

        elif choice == "10":
            print("Returning to main menu...")
            break   # 🔥 THIS is what exits the loop

        else:
            print("Invalid option. Try again.")

def production_forecast():
    print("\n--- EGG PRODUCTION FORECAST ---")

    if len(records) < 3:
        print("At least 3 records are required for a forecast.")
        return

    # Use the last 3 production values
    last_three = [record["eggs_produced"] for record in records[-3:]]

    forecast = sum(last_three) / len(last_three)

    print(f"Last 3 days production: {last_three}")
    print(f"Average of last 3 days: {forecast:.0f} eggs")
    print(f"Predicted production for next day: {forecast:.0f} eggs")

def feed_efficiency():
    print("\n--- FEED EFFICIENCY ANALYSIS ---")

    if not records:
        print("No records available.")
        return

    total_feed = sum(record["feed_used"] for record in records)
    total_eggs = sum(record["eggs_produced"] for record in records)

    if total_feed == 0:
        print("Feed used cannot be zero.")
        return

    efficiency = total_eggs / total_feed

    print(f"Total Feed Used      : {total_feed:.2f} kg")
    print(f"Total Eggs Produced  : {total_eggs}")
    print(f"Feed Efficiency      : {efficiency:.2f} eggs/kg")

    if efficiency >= 25:
        print("Performance          : 🟢 Excellent")
    elif efficiency >= 20:
        print("Performance          : 🟡 Good")
    else:
        print("Performance          : 🔴 Needs Improvement")

# ---------------- MAIN PROGRAM ----------------

current_user = login()

if current_user:
    load_records()

while True:
    print("\n=== SMART EGG PRODUCTION SYSTEM ===")
    print("1. Add Record")
    print("2. View Records")
    print("3. Summary")
    print("4. Search Record")
    print("5. Delete Record")
    print("6. Edit Record")
    print("7. Exit")
    print("8. Egg Production Trend")
    print("9. Farm Dashboard")
    print("10. Best & Worst Production Days")
    print("11. Best vs Worst Production Chart")
    print("12. Business Analytics") 
    print("13. Production Forecast")
    print("14. Feed Efficiency Analysis")
    print("15. Export to Excel")
    print("16. Export to PDF")
    print("17. User Management")

    choice = input("Choose an option: ")


    if choice == "1":
        add_record()

    elif choice == "2":
        view_records()

    elif choice == "3":
        summary()

    elif choice == "4":
        search_record()

    elif choice == "5":
        delete_record()

    elif choice == "6":
        edit_record()

    elif choice == "7":
        save_records()
        print("Goodbye!")
        break

    elif choice == "8":
        production_trend()

    elif choice == "9":
        show_kpis()
        production_rate_chart()

    elif choice == "10":
        best_worst_days()

    elif choice == "11":
        best_worst_chart()

    elif choice == "12":
        business_analytics()     

    elif choice == "13":
        production_forecast()

    elif choice == "14":
        feed_efficiency()
        
    elif choice == "15":
        export_to_excel()
        
    elif choice == "16":
        export_to_pdf()    
        
    elif choice == "17":
        user_management(current_user) 

    else:
        print("Invalid option. Please choose a number from 1 to 14.")
